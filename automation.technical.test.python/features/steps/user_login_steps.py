import openpyxl
from behave import when, then
from pages.user_login_page import UserLoginPage  # Importar la página de login

class UserLoginSteps:
    def __init__(self):
        self.file_path = 'userLogin.xlsx'  # Ruta al archivo Excel con los datos de login

    def get_credentials(self, row):
        """
        Esta función lee los datos del archivo Excel y devuelve la fila correspondiente de datos.
        :param row: Número de fila (basado en 0)
        :return: Una tupla con los datos: correo, contraseña
        """
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        
        # Leer los datos de la fila
        email = sheet.cell(row=row + 2, column=1).value  # Correo
        password = sheet.cell(row=row + 2, column=2).value  # Contraseña

        # Imprimir los valores leídos desde el archivo Excel
        print(f"Fila {row + 1}:")
        print(f"Email: {email}, Contraseña: {password}")

        # Retornar los datos leídos como una tupla
        return email, password


@when('El usuario ingresa al módulo de logueo de usuario')
def step_impl(context):
    context.page = UserLoginPage(context.driver)
    context.page.open_user_login()


@when('Ingresa el correo electronico desde el archivo userLogin.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1  # Convertir fila 1-based a 0-based
    print(f"Levantando datos de la fila {fila} desde Excel...")
    user_data = context.user_login_steps.get_credentials(row)

    # Verificar si los datos están siendo leídos correctamente
    if user_data is None:
        print(f"Error al leer los datos de la fila {fila}")
        return

    email = user_data[0]
    print(f"Correo leído: {email}")
    context.page.enter_email(email)


@when('Ingresa la contraseña desde el archivo userLogin.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1
    user_data = context.user_login_steps.get_credentials(row)
    password = user_data[1]
    print(f"Contraseña leída: {password}")
    context.page.enter_password(password)


@then('Confirmar login de usuario')
def step_impl(context):
    print("Confirmando el login del usuario...")
    context.page.confirm_login()
