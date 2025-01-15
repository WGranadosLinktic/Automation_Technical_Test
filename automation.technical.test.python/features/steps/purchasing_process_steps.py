import openpyxl
from behave import when, then
from pages.purchasing_process_page import PurchasingPage

class PurchasingProcessSteps:
    def __init__(self):
        self.file_path = 'customerData.xlsx'  # Ruta al archivo Excel con los datos del cliente

    def get_customer_data(self, row):
        """
        Esta función lee los datos del archivo Excel y devuelve la fila correspondiente de datos del cliente.
        :param row: Número de fila (basado en 0)
        :return: Una tupla con los datos: nombre, dirección, teléfono, etc.
        """
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active

        # Leer los datos de la fila
        first_name = sheet.cell(row=row + 2, column=1).value  # Nombre
        last_name = sheet.cell(row=row + 2, column=2).value  # Apellido
        email = sheet.cell(row=row + 2, column=3).value  # Correo electrónico
        phone = sheet.cell(row=row + 2, column=4).value  # Teléfono
        company = sheet.cell(row=row + 2, column=5).value  # Compañia
        address1 = sheet.cell(row=row + 2, column=6).value  # Dirección 1
        address2 = sheet.cell(row=row + 2, column=7).value  # Dirección 2
        city = sheet.cell(row=row + 2, column=8).value  # Ciudad
        postcode = sheet.cell(row=row + 2, column=9).value  # Código postal
        
        # Retornar los datos leídos como una tupla
        return first_name, last_name, email, phone, company, address1, address2, city, postcode

@when('El usuario selecciona un producto de la tienda')
def step_impl(context):
    context.page = PurchasingPage(context.driver)
    context.page.select_product()


@when('Agrega el producto al carrito de compras')
def step_impl(context):
    context.page.add_product_to_cart()


@when('Ingresa al carrito de compras')
def step_impl(context):
    context.page.view_cart()


@when('Inicia proceso de pago')
def step_impl(context):
    context.page.start_checkout_process()


@when('Diligencia los datos del cliente desde el archivo customerData.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1  # Convertir fila 1-based a 0-based
    print(f"Levantando datos de la fila {fila} desde Excel...")
    customer_data = context.purchasing_process_steps.get_customer_data(row)

    # Verificar si los datos están siendo leídos correctamente
    if customer_data is None:
        print(f"Error al leer los datos de la fila {fila}")
        return

    first_name, last_name, email, phone, company, address1, address2, city, postcode = customer_data
    context.page.enter_customer_details(first_name, last_name, email, phone, company, address1, address2, city, postcode)


@when('Selecciona el metodo de pago')
def step_impl(context):
    context.page.select_payment_method()


@then('Confirma compra')
def step_impl(context):
    context.page.confirm_purchase()