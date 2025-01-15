import openpyxl
from behave import when, then
from pages.user_register_page import UserRegisterPage

class UserRegisterSteps:
    def __init__(self):
        self.file_path = 'userRegister.xlsx'  # Ruta al archivo Excel con los datos de registro

    def get_credentials(self, row):
        """
        Esta función lee los datos del archivo Excel y devuelve la fila correspondiente de datos.
        :param row: Número de fila (basado en 0)
        :return: Una tupla con los datos: nombre, apellido, correo, teléfono, contraseña, confirmación de contraseña
        """
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        
        # Leer los datos de la fila
        first_name = sheet.cell(row=row + 2, column=1).value  # Nombre
        last_name = sheet.cell(row=row + 2, column=2).value   # Apellido
        email = sheet.cell(row=row + 2, column=3).value        # Email
        telephone = sheet.cell(row=row + 2, column=4).value    # Teléfono
        password = sheet.cell(row=row + 2, column=5).value     # Contraseña
        confirm_password = sheet.cell(row=row + 2, column=6).value  # Confirmación de contraseña

        # Imprimir los valores leídos desde el archivo Excel
        print(f"Fila {row + 1}:")
        print(f"Nombre: {first_name}, Apellido: {last_name}, Email: {email}, Teléfono: {telephone}, Contraseña: {password}, Confirmar Contraseña: {confirm_password}")

        # Retornar los datos leídos como una tupla
        return first_name, last_name, email, telephone, password, confirm_password


@when('El usuario ingresa al módulo de registro de usuario')
def step_impl(context):
    context.page = UserRegisterPage(context.driver)
    context.page.open_user_register()


@when('Ingresa el nombre desde el archivo userRegister.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1  # Convertir fila 1-based a 0-based
    print(f"Levantando datos de la fila {fila} desde Excel...")
    user_data = context.user_register_steps.get_credentials(row)

    # Verificar si los datos están siendo leídos correctamente
    if user_data is None:
        print(f"Error al leer los datos de la fila {fila}")
        return

    first_name = user_data[0]
    print(f"Nombre leido: {first_name}")
    context.page.enter_first_name(first_name)


@when('Ingresa el apellido desde el archivo userRegister.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1
    user_data = context.user_register_steps.get_credentials(row)
    last_name = user_data[1]
    print(f"Apellido leído: {last_name}")
    context.page.enter_last_name(last_name)


@when('Ingresa el correo electronico desde el archivo userRegister.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1
    user_data = context.user_register_steps.get_credentials(row)
    email = user_data[2]
    print(f"Correo leído: {email}")
    context.page.enter_email(email)


@when('Ingresa el telefono desde el archivo userRegister.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1
    user_data = context.user_register_steps.get_credentials(row)
    telephone = user_data[3]
    print(f"Teléfono leído: {telephone}")
    context.page.enter_telephone(telephone)


@when('Ingresa la contraseña desde el archivo userRegister.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1
    user_data = context.user_register_steps.get_credentials(row)
    password = user_data[4]
    print(f"Contraseña leída: {password}")
    context.page.enter_password(password)


@when('Ingresa nuevamente la contraseña desde el archivo userRegister.xlsx fila "{fila}"')
def step_impl(context, fila):
    row = int(fila) - 1
    user_data = context.user_register_steps.get_credentials(row)
    confirm_password = user_data[5]
    print(f"Confirmar contraseña leída: {confirm_password}")
    context.page.enter_confirm_password(confirm_password)


@then('Confirmar creación de usuario')
def step_impl(context):
    print("Confirmando la creación del usuario...")
    context.page.confirm_register()
